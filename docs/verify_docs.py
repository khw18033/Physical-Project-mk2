# -*- coding: utf-8 -*-
"""읽기순서가 가리키는 문서들의 내용 검증.

검사 축
  A. 읽기순서가 가리키는 문서가 실존하는가
  B. 각 문서가 본문에서 가리키는 파일 경로가 실존하는가
  C. 인용한 요구사항 ID가 시트에 실존하는가 (내 것 + 타 파트)
  D. 문서에 적힌 npm 스크립트가 package.json에 있는가
  E. 문서가 이름을 든 코드 심볼이 코드에 있는가
  F. 마크다운 코드펜스가 짝이 맞는가
"""
import io, os, re, json, sys
import openpyxl

# 저장소 루트에서 실행한다: python docs/verify_docs.py
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

DOCS = [
    "가시화_구현현황_260824.md",
    "작업계획_260824.md",
    "현황정리_260824.md",
    "구현계획서.md",
    "가시화_프로토타입_설명서.md",
    "요구사항정의서.md",
    "가시화_요구사항_쉬운설명.md",
    "백엔드_회신_260824.md",
    "docs/다이어그램_원본_mermaid.md",
    "web-dashboard/README.md",
    "unity-twin/README.md",
]

problems = []
notes = []

# 실존하지 않아도 정상인 경로 — 이유를 함께 남긴다.
ALLOWED_MISSING = {
    # 아카이브로 옮겼다는 사실 자체를 설명하는 문장에 등장한다.
    "web-dashboard/prototype.html": "_archive/prototype_260818.html 로 이동",
    "prototype.html": "_archive/prototype_260818.html 로 이동",
    # 백엔드가 토픽 방식을 고를 때 추가할 가상의 파일이다.
    "TopicTransport.ts": "미래 교체 시나리오의 예시 파일",
}


def p(msg):
    sys.stdout.write(msg + "\n")


# ── A. 존재 ──────────────────────────────────────────────────────────────────
p("■ A. 읽기순서가 가리키는 문서 존재")
order = io.open("현재상황_읽기순서.txt", encoding="utf-8").read()
for d in DOCS:
    exists = os.path.exists(d)
    named = os.path.basename(d) in order
    if not exists:
        problems.append("읽기순서가 가리키는 문서 없음: " + d)
    if not named:
        problems.append("DOCS 목록에 있으나 읽기순서에 이름 없음: " + d)
p("   문서 %d개 · 실존 %d개 · 읽기순서 언급 %d개"
  % (len(DOCS), sum(os.path.exists(d) for d in DOCS),
     sum(os.path.basename(d) in order for d in DOCS)))

# ── B. 본문이 가리키는 파일 경로 ────────────────────────────────────────────
p("\n■ B. 문서 본문이 가리키는 파일 경로")
# 백틱 안의 경로처럼 보이는 토큰
PATH_RE = re.compile(r"`([^`\s]+\.(?:md|ts|tsx|mjs|json|cs|png|pdf|docx|xlsx|html|txt|py))`")
checked = 0
for d in DOCS:
    text = io.open(d, encoding="utf-8").read()
    base = os.path.dirname(d)
    for m in PATH_RE.finditer(text):
        raw = m.group(1)
        if raw.startswith(("http", "//")):
            continue
        if raw in ALLOWED_MISSING:
            continue
        cands = [raw, os.path.join(base, raw)]
        # 저장소 어디에나 같은 이름이 있으면 통과로 본다(상대경로 표기 흔들림 허용)
        found = any(os.path.exists(c) for c in cands)
        if not found:
            hits = []
            for root, dirs, files in os.walk("."):
                dirs[:] = [x for x in dirs if x not in ("node_modules", ".git", "dist", "Library")]
                if os.path.basename(raw) in files:
                    hits.append(os.path.join(root, os.path.basename(raw)))
                    break
            found = bool(hits)
        checked += 1
        if not found:
            problems.append("%s → 존재하지 않는 경로 인용: %s" % (d, raw))
p("   경로 인용 %d건 검사" % checked)

# ── C. 요구사항 ID 유효성 ───────────────────────────────────────────────────
p("\n■ C. 인용한 요구사항 ID가 시트에 있는가")
wb = openpyxl.load_workbook("피지컬팀 프로젝트 mk2 요구사항 정의서.xlsx", data_only=True)
sheet_ids = {}
subclass = {}
for name in wb.sheetnames:
    s = wb[name]
    hdr = [(s.cell(1, c).value or "") for c in range(1, s.max_column + 1)]
    ic = hdr.index("ID") + 1 if "ID" in hdr else 1
    sc = hdr.index("소분류") + 1 if "소분류" in hdr else None
    for r in range(2, s.max_row + 1):
        v = s.cell(r, ic).value
        if not v:
            continue
        v = str(v).strip()
        if re.fullmatch(r"(HW|BE|AI|VZ|DT)-[A-Z]?-?\d{2}", v):
            sheet_ids[v] = name
            if sc:
                subclass[v] = str(s.cell(r, sc).value or "").strip()

ID_RE = re.compile(r"\b((?:HW|BE|AI|VZ)-[A-Z]-\d{2})\b")
cited = {}
for d in DOCS:
    text = io.open(d, encoding="utf-8").read()
    for i in set(ID_RE.findall(text)):
        cited.setdefault(i, []).append(d)
unknown = {i: v for i, v in cited.items() if i not in sheet_ids}
p("   시트 ID %d개 · 문서가 인용한 ID %d개" % (len(sheet_ids), len(cited)))
if unknown:
    for i, where in sorted(unknown.items()):
        problems.append("시트에 없는 ID 인용: %s (%s)" % (i, ", ".join(os.path.basename(w) for w in where)))
else:
    p("   시트에 없는 ID 인용: 없음")

# ── D. npm 스크립트 ─────────────────────────────────────────────────────────
p("\n■ D. 문서에 적힌 npm 스크립트")
pkg = json.load(io.open("web-dashboard/package.json", encoding="utf-8"))
scripts = set(pkg.get("scripts", {}))
NPM_RE = re.compile(r"npm(?:\.cmd)? run ([a-z:\-]+)")
found_scripts = set()
for d in DOCS:
    text = io.open(d, encoding="utf-8").read()
    for m in NPM_RE.finditer(text):
        name = m.group(1)
        found_scripts.add(name)
        if name not in scripts:
            problems.append("%s → package.json에 없는 스크립트: npm run %s" % (d, name))
p("   문서가 부르는 스크립트: %s" % ", ".join(sorted(found_scripts)))
p("   package.json 스크립트: %s" % ", ".join(sorted(scripts)))

# ── E. 코드 심볼 ────────────────────────────────────────────────────────────
p("\n■ E. 문서가 이름을 든 코드 심볼")
SYMBOLS = {
    "CACHE_POLICY": "web-dashboard/mock-gateway/config.ts",
    "cachedKeys": "web-dashboard/mock-gateway/hub.ts",
    "commandLatency": "web-dashboard/mock-gateway/controls.ts",
    "resolveAlignment": "web-dashboard/src/data/vision.ts",
    "referenceMissing": "web-dashboard/src/data/vision.ts",
    "EDGE_SILENCE_MS": "web-dashboard/src/data/vision.ts",
    "guardedMean": "web-dashboard/src/data/aggregation.ts",
}
for sym, path in SYMBOLS.items():
    body = io.open(path, encoding="utf-8").read() if os.path.exists(path) else ""
    mentioned = any(sym in io.open(d, encoding="utf-8").read() for d in DOCS)
    if mentioned and sym not in body:
        problems.append("문서가 언급한 심볼이 코드에 없음: %s (%s)" % (sym, path))
p("   검사 심볼 %d개" % len(SYMBOLS))

# ── F. 코드펜스 균형 ────────────────────────────────────────────────────────
p("\n■ F. 마크다운 코드펜스 균형")
for d in DOCS:
    text = io.open(d, encoding="utf-8").read()
    n = len(re.findall(r"^```", text, re.M))
    if n % 2 != 0:
        problems.append("%s → 코드펜스 홀수(%d개) — 닫히지 않은 블록" % (d, n))
p("   검사 완료")

# ── 결과 ────────────────────────────────────────────────────────────────────
p("\n" + "=" * 60)
if problems:
    p("발견 %d건" % len(problems))
    for x in problems:
        p("  - " + x)
else:
    p("A~F 전 항목 이상 없음")
for n in notes:
    p("  note: " + n)
